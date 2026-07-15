from sqlalchemy.orm import Session
from model.societe_leads import societeleads
from fastapi import HTTPException,Depends
from model.leads import Leads
from model.blacklistLeads import blacklistLeads
from fastapi.responses import StreamingResponse
from model.cleaning_leads import cleaningleads
from model.statistiqueLeads import StatisticLeads
from model.staging_leads import StagingLeads
from model.staging_import_history import StagingImportHistory
from model.steaging_applique import SteagingApplique
from sqlalchemy import or_,and_
from sqlalchemy import text
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import make_msgid
import csv
import io
import zipfile
from sqlalchemy.exc import SQLAlchemyError
from service import service as se
from database.db import get_db
import unicodedata
from util.util import NettoyerUnEmail
import re
import imaplib
import email
import time
import uuid
import socket
import threading
import dns.resolver
from concurrent.futures import ThreadPoolExecutor, as_completed
from database.db import SessionLocal
def _norm_company_key(v: str | None) -> str:
    if not v:
        return ""
    s = str(v).strip().lower()
    # enlever accents
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    # garder alphanum seulement
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def _norm_name_part(v: str | None) -> str:
    if not v:
        return ""
    s = str(v).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def _build_email(patterne: str, prenom: str, nom: str) -> str:
    # patterne propre à la société, ex: "{prenom}.{nom}@soprat.fr" ou "{n}{prenom}@soprat.fr"
    p = patterne or ""
    return (
        p.replace("{prenom}", prenom)
         .replace("{nom}", nom)
         .replace("{p}", (prenom or "")[:1])
         .replace("{n}", (nom or "")[:1])
    )

def SteagingAppliqueToSilver(db: Session, ids: list[int], pattern: str | None = None):
    """
    Déplace une sélection de staging_leads vers leads.
    Conditions:
    - société existe dans societe_leads (match par nom "normalisé")
    - on génère l'email si manquant (pattern + domaine/ext de la société)
    - on n'insère pas si email vide ou déjà existant (unique) dans leads
    """
    if not ids:
        return {"moved_to_silver": 0, "skipped": 0, "details": []}

    try:
        # pattern depuis app_settings si non fourni
        if not pattern:
            try:
                pattern = se.GetEmailPattern(db)
            except Exception:
                pattern = "{prenom}.{nom}@{domaine}.{extension}"

        # map des sociétés (nom normalisé -> patterne)
        companies = db.query(societeleads.nom, societeleads.patterne).all()
        company_map: dict[str, str] = {}
        for nom_c, patt in companies:
            k = _norm_company_key(nom_c)
            if not k:
                continue
            company_map[k] = str(patt or "").strip()

        leads = db.query(SteagingApplique).filter(SteagingApplique.id.in_(ids)).all()
        moved = 0
        skipped = 0
        deleted_already_in_silver = 0
        details: list[dict] = []
        moved_ids: list[int] = []
        deleted_ids: list[int] = []

        deleted_duplicates = 0

        def _delete_duplicates_for(lid: int, email_val: str, nom_val: str | None, prenom_val: str | None, societe_val: str | None):
            nonlocal deleted_duplicates
            try:
                # Priorité email si présent
                e = (email_val or "").strip().lower()
                if e and e not in ("nan", "none", "null"):
                    res = db.execute(text("""
                        DELETE FROM staging_leads
                        WHERE id <> :id
                          AND LOWER(TRIM(COALESCE(email,''))) = :email
                    """), {"id": lid, "email": e})
                    deleted_duplicates += int(res.rowcount or 0)
                    return
                # Sinon nom+prenom+societe
                n = _norm_name_part(nom_val)
                p = _norm_name_part(prenom_val)
                s = _norm_company_key(societe_val)
                if not n or not p or not s:
                    return
                rows = db.execute(text("""
                    SELECT id, nom, prenom, societe
                    FROM staging_leads
                    WHERE id <> :id
                """), {"id": lid}).mappings().all()
                to_delete = []
                for r in rows:
                    if _norm_name_part(r.get("nom")) == n and _norm_name_part(r.get("prenom")) == p and _norm_company_key(r.get("societe")) == s:
                        to_delete.append(r.get("id"))
                if to_delete:
                    res = db.execute(text("DELETE FROM staging_leads WHERE id = ANY(:ids)"), {"ids": to_delete})
                    deleted_duplicates += int(res.rowcount or 0)
            except Exception:
                return

        for l in leads:
            try:
                # Email vérifié comme non disponible -> on n'envoie pas vers Silver
                if str(getattr(l, "statu", "") or "").strip().lower() == "non disponible":
                    skipped += 1
                    details.append({"id": l.id, "reason": "email_non_disponible"})
                    continue

                soc_key = _norm_company_key(l.societe)
                if not soc_key or soc_key not in company_map:
                    skipped += 1
                    details.append({"id": l.id, "reason": "societe_not_found"})
                    continue

                patt = company_map[soc_key]
                if not patt:
                    skipped += 1
                    details.append({"id": l.id, "reason": "societe_pattern_missing"})
                    continue

                email = (l.email or "").strip()
                if not email or email.lower() in ("nan", "none", "null"):
                    prenom = _norm_name_part(l.prenom)
                    nom = _norm_name_part(l.nom)
                    if not prenom or not nom:
                        skipped += 1
                        details.append({"id": l.id, "reason": "name_missing"})
                        continue
                    email = _build_email(patt, prenom, nom).strip().lower()
                    email = NettoyerUnEmail(email) or email

                if not email or "@" not in email or "{" in email:
                    skipped += 1
                    details.append({"id": l.id, "reason": "email_invalid"})
                    continue

                # unique constraint: skip si existe déjà
                exists = db.query(Leads.id).filter(Leads.email == email).first()
                if exists:
                    # L'utilisateur veut que ce lead disparaisse de staging_leads
                    db.delete(l)
                    deleted_already_in_silver += 1
                    deleted_ids.append(int(l.id))
                    _delete_duplicates_for(l.id, email, l.nom, l.prenom, l.societe)
                    details.append({"id": l.id, "reason": "deleted_already_in_silver"})
                    continue

                obj = Leads(
                    email=email,
                    nom=l.nom,
                    prenom=l.prenom,
                    fonction=l.fonction,
                    societe=l.societe,
                    telephone=l.telephone,
                    linkedin=l.linkedin,
                    location=l.location,
                    statu=getattr(l, "statu", None),   # report du statut vérifié en Applique
                )
                db.add(obj)
                db.delete(l)
                moved += 1
                moved_ids.append(int(l.id))
                _delete_duplicates_for(l.id, email, l.nom, l.prenom, l.societe)
            except Exception:
                skipped += 1
                details.append({"id": getattr(l, "id", None), "reason": "error"})

        db.commit()
        return {
            "moved_to_silver": moved,
            "moved_ids": moved_ids,
            "deleted_already_in_silver": deleted_already_in_silver,
            "deleted_ids": deleted_ids,
            "deleted_duplicates": deleted_duplicates,
            "skipped": skipped,
            "details": details,
        }
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
def _filtre_completion(db: Session, operateur: str):
    """Filtre les leads sur la complétion calculée à la volée (non stockée)."""
    from service.service import sql_completion_expr
    return db.query(Leads).filter(text(f"{sql_completion_expr()} {operateur}"))

def GetAllLeads(db:Session):
    """Table unifiée : tous les leads. La complétion est calculée côté front."""
    return db.query(Leads).all()
def GetAllSilver(db:Session):
    """Vue rétro-compatible : les leads incomplets (< 100%)."""
    return _filtre_completion(db, "< 100").all()
def GetAllGold(db:Session):
    """Vue rétro-compatible : les leads complets (Gold, 100%)."""
    return _filtre_completion(db, "= 100").all()
def GetAllBlack(db:Session):
    return db.query(blacklistLeads).all()
def GetAllClean(db:Session):
    return db.query(cleaningleads).all()

def DeleteCleanByIds(db: Session, ids: list[int]):
    if not ids:
        return {"deleted": 0}
    try:
        res = db.execute(text("DELETE FROM cleaning_leads WHERE id = ANY(:ids)"), {"ids": ids})
        db.commit()
        return {"deleted": int(res.rowcount or 0)}
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
def GetAllStat(db: Session, userid: str | None = None, is_manager: bool = False):
    q = db.query(StatisticLeads)
    if not is_manager:
        if not userid:
            return []
        q = q.filter(StatisticLeads.iduser == userid)
    return q.order_by(StatisticLeads.created_at.desc()).all()
def GetAllStaging(db:Session):
    return db.query(StagingLeads).all()
def GetAllSteagingApplique(db: Session):
    return db.query(SteagingApplique).all()

def ExportDatabaseZip(db: Session, is_manager: bool):
    if not is_manager:
        raise HTTPException(status_code=403, detail="Accès refusé: manager seulement")

    # Assurer colonnes export utiles (compat rétro)
    try:
        db.execute(text("ALTER TABLE staging_import_history ADD COLUMN IF NOT EXISTS username TEXT"))
        db.commit()
    except Exception:
        db.rollback()

    # Liste blanche des tables à exporter
    tables = [
        "societe_leads",
        "import_leads",
        "staging_import_history",
        "cleaning_leads",
        "leads",
        "blacklist_leads",
        "staging_leads",
        "statistic_leads",
        "validation_rule",
        "token",
        "app_settings",
    ]

    def to_str(v):
        if v is None:
            return ""
        return str(v)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for table in tables:
            try:
                # Forcer l'ordre des colonnes pour staging_import_history afin de rendre "username" visible.
                if table == "staging_import_history":
                    rows = db.execute(text("""
                        SELECT
                            id,
                            filename,
                            COALESCE(username, '') AS username,
                            nom,
                            prenom,
                            email,
                            fonction,
                            societe,
                            telephone,
                            linkedin,
                            location,
                            imported_at
                        FROM staging_import_history
                    """)).mappings().all()
                else:
                    rows = db.execute(text(f"SELECT * FROM {table}")).mappings().all()
            except Exception:
                # table non existante dans certains environnements
                continue

            # Si table vide: ne pas exporter de fichier
            if not rows:
                continue

            csv_buf = io.StringIO()
            if rows:
                headers = list(rows[0].keys())
                writer = csv.DictWriter(csv_buf, fieldnames=headers, delimiter=";", lineterminator="\n")
                writer.writeheader()
                for r in rows:
                    writer.writerow({k: to_str(r.get(k)) for k in headers})

            zf.writestr(f"{table}.csv", csv_buf.getvalue().encode("utf-8-sig"))

    buf.seek(0)
    filename = f"db_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def CountLastImportAlreadyProcessedInApplique(db: Session, filename: str, userid: str, inserted_rows: int) -> int:
    """
    Compte combien de lignes du dernier import (history) existent déjà dans staging_leads.
    Matching prioritaire par email (non vide), sinon par (nom, prenom, societe).
    """
    if not filename or not userid or inserted_rows <= 0:
        return 0
    try:
        res = db.execute(text("""
            WITH last_import AS (
                SELECT id, nom, prenom, email, societe
                FROM staging_import_history
                WHERE filename = :filename
                  AND iduser = :userid
                ORDER BY id DESC
                LIMIT :lim
            ),
            normalized AS (
                SELECT
                    id,
                    LOWER(TRIM(COALESCE(nom, ''))) AS nom_n,
                    LOWER(TRIM(COALESCE(prenom, ''))) AS prenom_n,
                    LOWER(TRIM(COALESCE(email, ''))) AS email_n,
                    LOWER(TRIM(COALESCE(societe, ''))) AS societe_n
                FROM last_import
            )
            SELECT COUNT(*)
            FROM normalized li
            WHERE (
                (li.email_n <> '' AND li.email_n <> 'nan' AND EXISTS (
                    SELECT 1 FROM staging_leads sa
                    WHERE LOWER(TRIM(COALESCE(sa.email, ''))) = li.email_n
                ))
                OR
                ((li.email_n = '' OR li.email_n = 'nan') AND EXISTS (
                    SELECT 1 FROM staging_leads sa
                    WHERE LOWER(TRIM(COALESCE(sa.nom, ''))) = li.nom_n
                      AND LOWER(TRIM(COALESCE(sa.prenom, ''))) = li.prenom_n
                      AND LOWER(TRIM(COALESCE(sa.societe, ''))) = li.societe_n
                ))
            )
        """), {"filename": filename, "userid": userid, "lim": int(inserted_rows)}).scalar()
        return int(res or 0)
    except Exception:
        # Ne pas casser l'upload si ce check échoue
        return 0

def UpdateSilverEmail(db: Session, lead_id: int, email: str):
    try:
        cleaned = NettoyerUnEmail(email)
        if not cleaned:
            raise HTTPException(status_code=400, detail="Email invalide")

        lead = db.query(Leads).filter(Leads.id == lead_id).first()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead silver introuvable")

        lead.email = cleaned
        db.commit()
        db.refresh(lead)
        return {"message": "Email mis à jour", "email": lead.email}
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        # cas typique: contrainte unique sur email
        msg = str(e)
        if "UniqueViolation" in msg or "duplicate key" in msg or "unique constraint" in msg:
            raise HTTPException(status_code=409, detail="Email déjà utilisé")
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {msg}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
def UpdateLeadField(db: Session, lead_id: int, field: str, value):
    """Met à jour un champ d'un lead depuis le tableau.
    La complétion n'est pas stockée : le front la recalcule à partir des champs."""
    from service.service import COMPLETION_FIELDS
    try:
        if field not in COMPLETION_FIELDS:
            raise HTTPException(status_code=400, detail=f"Champ '{field}' non modifiable")

        lead = db.query(Leads).filter(Leads.id == lead_id).first()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead introuvable")

        # Une chaîne vide (ou parasite) = champ vidé -> NULL, la complétion baisse
        cleaned = str(value or "").strip()
        if cleaned.lower() in ("", "nan", "none", "null"):
            cleaned = None

        if field == "email" and cleaned:
            exists = (
                db.query(Leads.id)
                .filter(Leads.email == cleaned, Leads.id != lead_id)
                .first()
            )
            if exists:
                raise HTTPException(status_code=409, detail="Email déjà utilisé par un autre lead")

        setattr(lead, field, cleaned)
        db.commit()
        db.refresh(lead)

        return {
            "id": lead.id,
            "field": field,
            "value": cleaned,
        }
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        msg = str(e)
        if "UniqueViolation" in msg or "duplicate key" in msg or "unique constraint" in msg:
            raise HTTPException(status_code=409, detail="Email déjà utilisé")
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {msg}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")


def ClearBaseTable(db: Session, base: str):
    try:
        result = db.execute(text(f"DELETE FROM {base}"))
        db.commit()
        return {"staging_cleared_rows": result.rowcount}
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
def GetAllStagingImportHistory(db: Session, userid: str | None = None):
    def norm(v):
        s = (v or "").strip().lower()
        return "" if s == "nan" else s

    # Filtre user côté backend pour réduire fortement le volume.
    history_query = db.query(StagingImportHistory)
    if userid:
        history_query = history_query.filter(StagingImportHistory.iduser == userid)
    history_rows = history_query.order_by(StagingImportHistory.imported_at.desc()).all()

    # Sets de matching en mémoire (beaucoup plus rapide que CASE/EXISTS corrélé).
    # Silver/Gold sont la même table `leads` : la destination se déduit de la
    # complétion, calculée à la volée (elle n'est pas stockée).
    _gold_q = _filtre_completion(db, "= 100")
    _silver_q = _filtre_completion(db, "< 100")
    gold_emails = {norm(l.email) for l in _gold_q.all() if norm(l.email)}
    silver_emails = {norm(l.email) for l in _silver_q.all() if norm(l.email)}
    clean_emails = {norm(x[0]) for x in db.query(cleaningleads.email).all() if norm(x[0])}
    blacklist_emails = {norm(x[0]) for x in db.query(blacklistLeads.email).all() if norm(x[0])}

    gold_keys = {(norm(l.nom), norm(l.prenom), norm(l.societe)) for l in _gold_q.all()}
    silver_keys = {(norm(l.nom), norm(l.prenom), norm(l.societe)) for l in _silver_q.all()}
    clean_keys = {(norm(r.nom), norm(r.prenom), norm(r.societe)) for r in db.query(cleaningleads.nom, cleaningleads.prenom, cleaningleads.societe).all()}

    enriched = []
    for h in history_rows:
        email_key = norm(h.email)
        tuple_key = (norm(h.nom), norm(h.prenom), norm(h.societe))

        # Règle: éliminer les leads blacklisted de l'historique
        if email_key and email_key in blacklist_emails:
            continue

        if email_key:
            if email_key in gold_emails:
                destination = "gold"
            elif email_key in silver_emails:
                destination = "silver"
            elif email_key in clean_emails:
                destination = "clean"
            else:
                destination = "staging"
        else:
            if tuple_key in gold_keys:
                destination = "gold"
            elif tuple_key in silver_keys:
                destination = "silver"
            elif tuple_key in clean_keys:
                destination = "clean"
            else:
                destination = "staging"

        enriched.append({
            "id": h.id,
            "filename": h.filename,
            "iduser": h.iduser,
            "nom": h.nom,
            "prenom": h.prenom,
            "email": h.email,
            "fonction": h.fonction,
            "societe": h.societe,
            "telephone": h.telephone,
            "linkedin": h.linkedin,
            "location": h.location,
            "imported_at": h.imported_at,
            "destination": destination,
        })

    return enriched
def DownloadProdLeadCSV(types:str,db: Session, ids: list[int] | None = None):
    try:
        # 1️⃣ Charger les données
        if ids:
            # Sélection explicite envoyée par le front (ex: les leads à 100%)
            leads = db.query(Leads).filter(Leads.id.in_(ids)).all()
        elif(types=="silver"):
            leads = _filtre_completion(db, "< 100").all()
        elif(types=="gold"):
            leads = _filtre_completion(db, "= 100").all()
        else:
            leads = db.query(Leads).all()

        
        if not leads:
            raise HTTPException(status_code=404, detail="Aucun lead à télécharger")

        # 2️⃣ Créer le buffer avec BOM UTF-8 (pour que Excel reconnaisse l'UTF-8)
        output = io.StringIO()
        output.write('\ufeff')  # BOM pour UTF-8
        
        # 3️⃣ Créer le writer CSV
        writer = csv.writer(
            output, 
            delimiter=';',           # Point-virgule pour Excel français
            quoting=csv.QUOTE_MINIMAL,
            lineterminator='\n'
        )
        
        # 4️⃣ En-têtes
        writer.writerow([
            "Nom",
            "Prénom",
            "Email",
            "Fonction",
            "Société",
            "Téléphone",
            "LinkedIn",
            "Location"
        ])
        
        # 5️⃣ Données (gérer les None)
        for lead in leads:
            writer.writerow([
                lead.nom or "",
                lead.prenom or "",
                lead.email or "",
                lead.fonction or "",
                lead.societe or "",
                lead.telephone or "",
                lead.linkedin or "",
                lead.location or ""
            ])
        
        output.seek(0)
        
        # 6️⃣ Nom du fichier avec timestamp
        filename = f"leads_silver_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # 7️⃣ Retourner le fichier
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du CSV : {str(e)}")
def DownloadLeadXlsx(types:str,db: Session, ids: list[int] | None = None):
    try:
        if ids:
            # Sélection explicite envoyée par le front (ex: les leads à 100%)
            leads = db.query(Leads).filter(Leads.id.in_(ids)).all()
        elif(types=="silver"):
            leads = _filtre_completion(db, "< 100").all()
        elif(types=="gold"):
            leads = _filtre_completion(db, "= 100").all()
        else:
            leads = db.query(Leads).all()
        
        if not leads:
            raise HTTPException(status_code=404, detail="Aucun lead à télécharger")

        # 2️⃣ Créer le workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Leads Silver"
        
        # 3️⃣ En-têtes
        headers = ["Nom", "Prénom", "Email", "Fonction", "Société", "Téléphone", "LinkedIn", "Location"]
        sheet.append(headers)
        
        # 4️⃣ Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 5️⃣ Ajouter les données
        for lead in leads:
            sheet.append([
                lead.nom,
                lead.prenom,
                lead.email,
                lead.fonction,
                lead.societe,
                lead.telephone,
                lead.linkedin,
                lead.location
            ])
        
        # 6️⃣ Auto-ajuster la largeur des colonnes
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Largeur avec marge
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 7️⃣ Bordures pour toutes les cellules
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Données (pas en-têtes)
                    cell.alignment = Alignment(vertical="center")
        
        # 8️⃣ Figer la première ligne
        sheet.freeze_panes = "A2"
        
        # 9️⃣ Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 🔟 Retourner le fichier
        filename = f"leads_silver_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du fichier : {str(e)}")

def _get_latest_import_rows_for_user(db: Session, userid: str):
    latest = (
        db.query(StagingImportHistory.filename)
        .filter(StagingImportHistory.iduser == userid)
        .order_by(StagingImportHistory.imported_at.desc())
        .first()
    )
    if not latest or not latest[0]:
        return None, []

    latest_filename = latest[0]
    rows = (
        db.query(StagingImportHistory)
        .filter(
            StagingImportHistory.iduser == userid,
            StagingImportHistory.filename == latest_filename
        )
        .order_by(StagingImportHistory.imported_at.desc(), StagingImportHistory.id.desc())
        .all()
    )

    # Règle export: éliminer les leads blacklisted
    blacklist_emails = {(e or "").strip().lower() for (e,) in db.query(blacklistLeads.email).all() if (e or "").strip().lower() not in ("", "nan")}
    if blacklist_emails:
        rows = [r for r in rows if ((r.email or "").strip().lower() not in blacklist_emails)]
    return latest_filename, rows

def DownloadLatestStagingImportCSV(db: Session, userid: str):
    try:
        latest_filename, rows = _get_latest_import_rows_for_user(db, userid)
        if not rows:
            raise HTTPException(status_code=404, detail="Aucun import trouvé pour cet utilisateur")

        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(
            output,
            delimiter=';',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator='\n'
        )

        writer.writerow(["Nom", "Prénom", "Email", "Fonction", "Société", "Téléphone", "LinkedIn", "Location"])
        for lead in rows:
            writer.writerow([
                lead.nom or "",
                lead.prenom or "",
                lead.email or "",
                lead.fonction or "",
                lead.societe or "",
                lead.telephone or "",
                lead.linkedin or "",
                lead.location or "",
            ])

        output.seek(0)
        filename = f"staging_last_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Import-Filename": latest_filename,
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export dernier import CSV : {str(e)}")

def DownloadLatestStagingImportXlsx(db: Session, userid: str):
    try:
        latest_filename, rows = _get_latest_import_rows_for_user(db, userid)
        if not rows:
            raise HTTPException(status_code=404, detail="Aucun import trouvé pour cet utilisateur")

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Dernier import staging"

        headers = ["Nom", "Prénom", "Email", "Fonction", "Société", "Téléphone", "LinkedIn", "Location"]
        sheet.append(headers)

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for lead in rows:
            sheet.append([
                lead.nom or "",
                lead.prenom or "",
                lead.email or "",
                lead.fonction or "",
                lead.societe or "",
                lead.telephone or "",
                lead.linkedin or "",
                lead.location or "",
            ])

        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="center")

        sheet.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"staging_last_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Import-Filename": latest_filename,
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export dernier import XLSX : {str(e)}")
def ToBlack(id:int,eliminer:str,db:Session):
    result =db.query(Leads).filter(Leads.id==id).first()
    if (not result):
        raise HTTPException(
               status_code=404,
               detail='Leads non trouvè'
        )
    print(result.nom)
    blocklead=blacklistLeads(
                    id=result.id,
                    nom=result.nom,
                    prenom= result.prenom,
                    email=result.email,
                    fonction= result.fonction,
                    societe= result.societe,
                    telephone=result.telephone,
                    linkedin= result.linkedin,
                    location=result.location,
                    eliminer=eliminer
                )
    db.add(blocklead)
    db.delete(result)
    db.commit()
    return {
            "message": "Le leads a èté blocque avec succeè"
        }
    
def StagingToSilver(db: Session,base:str):
    try:
        # 1️⃣ INSERT INTO leads depuis staging (évite les doublons)
        result = db.execute(text(f"""
            INSERT INTO leads (nom, prenom, email, fonction, societe, telephone, linkedin, location)
            SELECT DISTINCT ON (email) 
                nom, prenom, email, fonction, societe, telephone, linkedin, location
            FROM {base}
            WHERE email IS NOT NULL 
              AND email != '' 
              AND email != 'nan'
              AND nom IS NOT NULL 
              AND nom != '' 
              AND nom != 'nan'
              AND prenom IS NOT NULL 
              AND prenom != '' 
              AND prenom != 'nan'
              AND societe IS NOT NULL 
              AND societe != '' 
              AND societe != 'nan'
              AND (
                  fonction IS NULL OR fonction = '' OR fonction = 'nan'
                  OR telephone IS NULL OR telephone = '' OR telephone = 'nan'
                  OR linkedin IS NULL OR linkedin = '' OR linkedin = 'nan'
              )
              AND NOT EXISTS (
                  SELECT 1 FROM leads s WHERE s.email = {base}.email
              )
            ORDER BY email, id
        """))
        
        moved_count = result.rowcount
        
        # 2️⃣ DELETE depuis staging (ceux qui ont été déplacés + doublons internes)
        db.execute(text(f"""
            DELETE FROM {base}
            WHERE email IS NOT NULL 
              AND email != '' 
              AND email != 'nan'
              AND nom IS NOT NULL 
              AND nom != '' 
              AND nom != 'nan'
              AND prenom IS NOT NULL 
              AND prenom != '' 
              AND prenom != 'nan'
              AND societe IS NOT NULL 
              AND societe != '' 
              AND societe != 'nan'
              AND (
                  fonction IS NULL OR fonction = '' OR fonction = 'nan'
                  OR telephone IS NULL OR telephone = '' OR telephone = 'nan'
                  OR linkedin IS NULL OR linkedin = '' OR linkedin = 'nan'
              )
        """))
        
        db.commit()
        
        print(f"✅ {moved_count} leads déplacés vers Silver")
        return {"moved_to_silver": moved_count}

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
def MoveIncompleteToClean(db: Session, base: str = "import_leads"):
    """
    Déplace vers cleaning_leads (À corriger) les contacts trop incomplets :
      - email vide ET société vide (impossible de générer/vérifier un email), OU
      - au moins 3 champs vides parmi {nom, prénom, email, société}.
    Les retire de {base}.
    """
    def _empty(col):
        return f"LOWER(TRIM(COALESCE({col}, ''))) IN ('', 'nan', 'none', 'null')"

    def _cond(pfx=""):
        e, s = _empty(pfx + "email"), _empty(pfx + "societe")
        n, p = _empty(pfx + "nom"), _empty(pfx + "prenom")
        cnt = (f"((CASE WHEN {n} THEN 1 ELSE 0 END) + (CASE WHEN {p} THEN 1 ELSE 0 END)"
               f" + (CASE WHEN {e} THEN 1 ELSE 0 END) + (CASE WHEN {s} THEN 1 ELSE 0 END))")
        return f"(({e} AND {s}) OR {cnt} >= 3)"

    try:
        result = db.execute(text(f"""
            INSERT INTO cleaning_leads (nom, prenom, email, fonction, societe, telephone, linkedin, location)
            SELECT sl.nom, sl.prenom, sl.email, sl.fonction, sl.societe, sl.telephone, sl.linkedin, sl.location
            FROM {base} sl
            WHERE {_cond('sl.')}
              AND NOT EXISTS (
                    SELECT 1 FROM cleaning_leads cl
                    WHERE COALESCE(cl.nom, '') = COALESCE(sl.nom, '')
                      AND COALESCE(cl.prenom, '') = COALESCE(sl.prenom, '')
                      AND COALESCE(cl.email, '') = COALESCE(sl.email, '')
                      AND COALESCE(cl.societe, '') = COALESCE(sl.societe, '')
                      AND COALESCE(cl.telephone, '') = COALESCE(sl.telephone, '')
                      AND COALESCE(cl.linkedin, '') = COALESCE(sl.linkedin, '')
                      AND COALESCE(cl.location, '') = COALESCE(sl.location, '')
              )
        """))
        moved = result.rowcount or 0
        db.execute(text(f"DELETE FROM {base} WHERE {_cond('')}"))
        db.commit()
        print(f"✅ {moved} contacts incomplets -> Clean")
        return {"moved_to_clean": int(moved)}
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")


def StagingToClean(db: Session):
    try:
        # 1️⃣ Inserer les leads dont nom ou prenom reste vide
        # (la complétion depuis email est déjà tentée avant cet appel)
        result = db.execute(text("""
            INSERT INTO cleaning_leads (nom, prenom, email, fonction, societe, telephone, linkedin, location)
            SELECT
                nom, prenom, email, fonction, societe, telephone, linkedin, location
            FROM import_leads sl
            WHERE (
                    LOWER(TRIM(COALESCE(sl.nom, ''))) IN ('', 'nan')
                    OR LOWER(TRIM(COALESCE(sl.prenom, ''))) IN ('', 'nan')
                  )
              AND NOT EXISTS (
                   SELECT 1
                   FROM cleaning_leads cl
                   WHERE COALESCE(cl.nom, '') = COALESCE(sl.nom, '')
                     AND COALESCE(cl.prenom, '') = COALESCE(sl.prenom, '')
                     AND COALESCE(cl.email, '') = COALESCE(sl.email, '')
                     AND COALESCE(cl.fonction, '') = COALESCE(sl.fonction, '')
                     AND COALESCE(cl.societe, '') = COALESCE(sl.societe, '')
                     AND COALESCE(cl.telephone, '') = COALESCE(sl.telephone, '')
                     AND COALESCE(cl.linkedin, '') = COALESCE(sl.linkedin, '')
                     AND COALESCE(cl.location, '') = COALESCE(sl.location, '')
              )
        """))
        
        moved_count = result.rowcount
        
        # Supprimer uniquement les leads de staging qui matchent la règle clean
        db.execute(text("""
            DELETE FROM import_leads
            WHERE (
                    LOWER(TRIM(COALESCE(nom, ''))) IN ('', 'nan')
                    OR LOWER(TRIM(COALESCE(prenom, ''))) IN ('', 'nan')
                  )
        """))
        clean = se.SupprimerDoublonsMemetABLE(db, "cleaning_leads")
        net_moved = moved_count - clean["duplicates_deleted"]
        db.commit()

        
        print(f"✅ {net_moved} leads déplacés vers Cleaning")
        return {"moved_to_clean": net_moved}

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")

def StagingToSteagingApplique(db: Session, base: str):
    try:
        result = db.execute(text(f"""
            INSERT INTO staging_leads (nom, prenom, email, fonction, societe, telephone, linkedin, location)
            SELECT nom, prenom, email, fonction, societe, telephone, linkedin, location
            FROM {base}
        """))

        moved_count = result.rowcount

        db.execute(text(f"DELETE FROM {base}"))
        db.commit()

        return {"moved_to_steaging_applique": moved_count}
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")

def StagingToGold(db: Session,base:str):
    try:
        
        #  INSERT INTO leads depuis staging (évite les doublons)
        result = db.execute(text(f"""
            INSERT INTO leads (nom, prenom, email, fonction, societe, telephone, linkedin, location)
            SELECT DISTINCT ON (email) 
                nom, prenom, email, fonction, societe, telephone, linkedin, location
            FROM {base}
            WHERE email IS NOT NULL 
              AND email != '' 
              AND email != 'nan'
              AND nom IS NOT NULL 
              AND nom != '' 
              AND nom != 'nan'
              AND prenom IS NOT NULL 
              AND prenom != '' 
              AND prenom != 'nan'
              AND societe IS NOT NULL 
              AND societe != '' 
              AND societe != 'nan'
              AND fonction IS NOT NULL 
              AND fonction != '' 
              AND fonction != 'nan'
              AND telephone IS NOT NULL 
              AND telephone != '' 
              AND telephone != 'nan'
              AND linkedin IS NOT NULL 
              AND linkedin != '' 
              AND linkedin != 'nan'
              AND location IS NOT NULL
              AND location != ''
              AND location != 'nan'
              AND NOT EXISTS (
                  SELECT 1 FROM leads g WHERE g.email = {base}.email
              )
            ORDER BY email, id
        """))
        print(result)
        moved_count = result.rowcount
        
        #  DELETE depuis staging (ceux qui ont été déplacés + doublons internes)
        db.execute(text(f"""
            DELETE FROM {base}
            WHERE email IS NOT NULL 
              AND email != '' 
              AND email != 'nan'
              AND nom IS NOT NULL 
              AND nom != '' 
              AND nom != 'nan'
              AND prenom IS NOT NULL 
              AND prenom != '' 
              AND prenom != 'nan'
              AND societe IS NOT NULL 
              AND societe != '' 
              AND societe != 'nan'
              AND fonction IS NOT NULL 
              AND fonction != '' 
              AND fonction != 'nan'
              AND telephone IS NOT NULL 
              AND telephone != '' 
              AND telephone != 'nan'
              AND linkedin IS NOT NULL 
              AND linkedin != '' 
              AND linkedin != 'nan'
              AND location IS NOT NULL
              AND location != ''
              AND location != 'nan'
        """))
        
        db.commit()
        
        print(f"✅ {moved_count} leads déplacés vers Gold")
        return {"moved_to_gold": moved_count}

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
        


def CompleteSocieteFromEmail(db: Session,base:str):
    try:        
        # 1) Compléter societe depuis la table societe_leads à partir du domaine email
        result1 = db.execute(text(f"""
            UPDATE {base}
            SET societe = s.nom
            FROM societe_leads s
            WHERE ({base}.societe IS NULL OR {base}.societe = '' OR LOWER({base}.societe) = 'nan')
              AND {base}.email IS NOT NULL
              AND {base}.email != ''
              AND LOWER({base}.email) != 'nan'
              AND {base}.email LIKE '%@%.%'
              AND s.patterne IS NOT NULL
              AND s.patterne != ''
              AND LOWER(TRIM(SPLIT_PART(s.patterne, '@', 2))) = LOWER(TRIM(SPLIT_PART({base}.email, '@', 2)))
        """))

        # 2) Fallback: déduire la société directement du domaine email (sans table societe_leads)
        # Ex: user@mail.sqli.com -> sqli ; user@partoo.com -> partoo ; user@my-company.fr -> my company
        result2 = db.execute(text(f"""
            UPDATE {base}
            SET societe = INITCAP(
                REPLACE(
                    SPLIT_PART(
                        REGEXP_REPLACE(SPLIT_PART({base}.email, '@', 2), '\\.[^.]+$', ''),
                        '.',
                        -1
                    ),
                    '-',
                    ' '
                )
            )
            WHERE ({base}.societe IS NULL OR {base}.societe = '' OR LOWER({base}.societe) = 'nan')
              AND {base}.email IS NOT NULL
              AND {base}.email != ''
              AND LOWER({base}.email) != 'nan'
              AND {base}.email LIKE '%@%.%'
        """))
        
        db.commit()
        count = (result1.rowcount if hasattr(result1, "rowcount") else 0) + (result2.rowcount if hasattr(result2, "rowcount") else 0)
        
        print(f"✅ {count} sociétés complétées depuis les emails")
        return {"societe_completed": count}

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")



def CompleteNomPrenomFromEmail(db: Session,base:str):
    try:
        
        # UPDATE avec extraction du nom et prénom en SQL pur (PostgreSQL)
        result = db.execute(text(f"""
            UPDATE {base}
            SET 
                prenom = CASE 
                    WHEN (prenom IS NULL OR prenom = '') 
                         AND POSITION('.' IN SPLIT_PART(email, '@', 1)) > 0
                    THEN INITCAP(SPLIT_PART(SPLIT_PART(email, '@', 1), '.', 1))
                    ELSE prenom
                END,
                nom = CASE 
                    WHEN (nom IS NULL OR nom = '') 
                         AND POSITION('.' IN SPLIT_PART(email, '@', 1)) > 0
                    THEN INITCAP(SPLIT_PART(SPLIT_PART(email, '@', 1), '.', 2))
                    ELSE nom
                END
            WHERE email IS NOT NULL 
              AND email != ''
              AND email LIKE '%@%'
              AND POSITION('.' IN SPLIT_PART(email, '@', 1)) > 0
              AND (
                  (prenom IS NULL OR prenom = '') 
                  OR (nom IS NULL OR nom = '')
              )
        """))
        
        db.commit()
        count = result.rowcount
        
        print(f"✅ {count} noms/prénoms complétés depuis les emails")
        return {"nom_prenom_completed": count}

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")


def SilverToGold(db:Session,id:int):
    """Fusion Silver+Gold : il n'y a plus de déplacement entre tables.
    La complétion est calculée à la volée ; le lead est Gold si elle vaut 100%."""
    from service.service import sql_completion_expr
    try:
            lead = db.query(Leads).filter(Leads.id == id).first()
            if lead is None:
                raise HTTPException(status_code=404, detail="Lead introuvable")

            completion = int(db.execute(
                text(f"SELECT {sql_completion_expr()} FROM leads WHERE id = :id"),
                {"id": id},
            ).scalar() or 0)

            if completion < 100:
                raise HTTPException(
                    status_code=400,
                    detail=f"Lead incomplet ({completion}%) → impossible de passer en GOLD",
                )
            return {
                "message": "Lead Gold (100% complété)",
                "completion": completion,
                "gold": True,
            }

    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        print(str(e))
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")
regions_villes = {
    "Ile-de-France": {
        "villes": [
            "Paris", "Versailles", "Boulogne-Billancourt", "Saint-Denis", "Nanterre", "Creteil",
            "Argenteuil", "Montreuil", "Vitry-sur-Seine", "Aulnay-sous-Bois", "Colombes",
            "Asnieres-sur-Seine", "Courbevoie", "Rueil-Malmaison", "Saint-Maur-des-Fosses",
            "Champigny-sur-Marne", "Aubervilliers", "Vincennes", "Clichy", "Sceaux",
            "Ivry-sur-Seine", "Levallois-Perret", "Neuilly-sur-Seine", "Antony", "Noisy-le-Grand",
            "Massy", "Evry", "Corbeil-Essonnes", "Melun", "Meaux", "Cergy", "Pontoise",
            "Saint-Germain-en-Laye", "Poissy", "Mantes-la-Jolie", "Sarcelles",
            "Garges-les-Gonesse", "Drancy", "Bondy", "Pantin", "Bobigny", "Epinay-sur-Seine",
            "Gennevilliers", "Sartrouville", "Fontenay-sous-Bois", "Maisons-Alfort",
            "Issy-les-Moulineaux", "Montrouge", "Clamart", "Malakoff", "Bagneux", "Chatillon",
            "Cachan", "Arcueil", "Gentilly", "Romainville", "Bagnolet", "Rosny-sous-Bois",
            "Noisy-le-Sec", "Stains", "Tremblay-en-France", "Sevran", "Villepinte",
            "Puteaux", "Suresnes", "Houilles", "Bezons", "Franconville", "Ermont",
            "Enghien-les-Bains", "Villeneuve-Saint-Georges", "Orly", "Choisy-le-Roi",
            "Alfortville", "Villejuif", "Fresnes", "Velizy-Villacoublay", "Meudon",
            "Chaville", "Chatou", "Maisons-Laffitte", "Conflans-Sainte-Honorine",
            "Chelles", "Lagny-sur-Marne", "Fontainebleau", "Provins", "Coulommiers",
            "Ris-Orangis", "Viry-Chatillon", "Athis-Mons", "Juvisy-sur-Orge",
            "Savigny-sur-Orge", "Sainte-Genevieve-des-Bois", "Palaiseau", "Les Ulis",
            "Gif-sur-Yvette", "Orsay", "Rambouillet", "Plaisir", "Guyancourt",
            "Montigny-le-Bretonneux", "Trappes", "Elancourt", "Torcy", "Noisiel",
            "Dammarie-les-Lys", "Combs-la-Ville", "Savigny-le-Temple", "Lieusaint",
            "Brie-Comte-Robert", "Ozoir-la-Ferriere", "Villeneuve-le-Roi", "Longjumeau",
            "Morsang-sur-Orge", "Brunoy", "Yerres", "Montgeron", "Draveil", "Vigneux-sur-Seine"
        ]
    },
    "Auvergne-Rhone-Alpes": {
        "villes": [
            "Lyon", "Grenoble", "Clermont-Ferrand", "Saint-Etienne", "Villeurbanne", "Annecy",
            "Aubiere", "Chambery", "Valence", "Bourg-en-Bresse", "Roanne", "Thonon-les-Bains",
            "Annemasse", "Oyonnax", "Montlucon", "Vichy", "Moulins", "Aurillac", "Issoire",
            "Riom", "Thiers", "Ambert", "Brioude", "Cournon-d-Auvergne", "Cebazat",
            "Beaumont", "Chamalières", "Royat", "Pontarlier", "Romans-sur-Isere",
            "Vienne", "Bourgoin-Jallieu", "Givors", "Caluire-et-Cuire", "Decines-Charpieu",
            "Venissieux", "Bron", "Saint-Priest", "Meyzieu", "Rillieux-la-Pape",
            "Oullins", "Pierre-Benite", "Saint-Fons", "Feyzin", "Mions", "Chassieu",
            "Genas", "Jonage", "Pusignan", "Craponne", "Tassin-la-Demi-Lune",
            "Ecully", "Sainte-Foy-les-Lyon", "Francheville", "Charbonnieres-les-Bains",
            "Marcy-l'Etoile", "La Tour-de-Salvagny", "Dardilly", "Limonest", "Lissieu",
            "Neuville-sur-Saone", "Fontaines-sur-Saone", "Collonges-au-Mont-d-Or",
            "Couzon-au-Mont-d-Or", "Rochetaillee-sur-Saone", "Albigny-sur-Saone",
            "Genay", "Massieux", "Trevoux", "Reyrieux", "Beynost", "Miribel",
            "Saint-Maurice-de-Beynost", "Neyron", "Montluel", "Perouges",
            "Meximieux", "Lagnieu", "Ambronay", "Bellegarde-sur-Valserine",
            "Ferney-Voltaire", "Gex", "Saint-Genis-Pouilly", "Divonne-les-Bains",
            "Prevessin-Moens", "Thoiry", "Ornex", "Cessy", "Segny", "Echenevex",
            "Crozet", "Mijoux", "Chevry", "Sergy", "Challex", "Peron", "Vuache",
            "Charvonnex", "Seynod", "Cran-Gevrier", "Meythet", "Pringy", "Metz-Tessy",
            "Poisy", "Epagny", "Sillingy", "Argonay", "Villaz", "La Balme-de-Sillingy",
            "Rumilly", "Alby-sur-Cheran", "Ugine", "Faverges", "Doussard",
            "Saint-Jorioz", "Sevrier", "Duingt", "Talloires", "Menthon-Saint-Bernard"
        ]
    },
    "Provence-Alpes-Cote d'Azur": {
        "villes": [
            "Marseille", "Nice", "Toulon", "Aix-en-Provence", "Avignon", "Cannes",
            "Antibes", "Frejus", "La Seyne-sur-Mer", "Hyeres", "Arles", "Salon-de-Provence",
            "Gap", "Martigues", "Aubagne", "Draguignan", "Cagnes-sur-Mer", "Grasse",
            "Digne-les-Bains", "Manosque", "Brignoles", "Sanary-sur-Mer", "Ollioules",
            "La Garde", "Istres", "Vitrolles", "Miramas", "Port-de-Bouc", "Marignane",
            "Gardanne", "Pertuis", "Trets", "Rousset", "Chateauneuf-le-Rouge",
            "Venelles", "Meyreuil", "Bouc-Bel-Air", "Cabries", "Les Pennes-Mirabeau",
            "Septemes-les-Vallons", "Allauch", "Plan-de-Cuques", "La Destrousse",
            "Auriol", "Saint-Zacharie", "Nans-les-Pins", "Signes", "La Cadiere-d-Azur",
            "Bandol", "Six-Fours-les-Plages", "Le Pradet", "Carqueiranne", "Solliès-Pont",
            "La Valette-du-Var", "La Garde", "Cuers", "Pierrefeu-du-Var",
            "Collobrieres", "Grimaud", "Saint-Tropez", "Ramatuelle", "Gassin",
            "Cogolin", "La Croix-Valmer", "Cavalaire-sur-Mer", "Le Lavandou",
            "Bormes-les-Mimosas", "La Londe-les-Maures", "Pierrevert", "Vinon-sur-Verdon",
            "Greoux-les-Bains", "Valensole", "Moustiers-Sainte-Marie", "Riez",
            "Castellane", "Annot", "Entrevaux", "Puget-Theniers", "Guillaumes",
            "Valberg", "Beuil", "Isola", "Saint-Martin-Vesubie", "Tende", "Breil-sur-Roya",
            "Menton", "Roquebrune-Cap-Martin", "Monaco", "Beausoleil", "La Turbie",
            "Eze", "Villefranche-sur-Mer", "Beaulieu-sur-Mer", "Saint-Jean-Cap-Ferrat",
            "Roquette-sur-Siagne", "Mougins", "Vallauris", "Golfe-Juan", "Juan-les-Pins",
            "Biot", "Valbonne", "Sophia-Antipolis", "Mouans-Sartoux", "Peymeinade",
            "Saint-Cezaire-sur-Siagne", "Tanneron", "Mandelieu-la-Napoule", "Theoul-sur-Mer",
            "Le Cannet", "Ranguin", "Le Bosquet", "Rocheville"
        ]
    },
    "Occitanie": {
        "villes": [
            "Toulouse", "Montpellier", "Nimes", "Perpignan", "Beziers", "Albi",
            "Carcassonne", "Tarbes", "Castres", "Narbonne", "Sete", "Lunel",
            "Agde", "Mende", "Rodez", "Auch", "Foix", "Cahors", "Millau",
            "Montauban", "Alès", "Bagnols-sur-Ceze", "Beaucaire", "Saint-Gilles",
            "Lunel", "Mauguio", "Lattes", "Perols", "Palavas-les-Flots",
            "Le Grau-du-Roi", "Aigues-Mortes", "La Grande-Motte",
            "Castelnau-le-Lez", "Clapiers", "Jacou", "Vendargues", "Baillargues",
            "Lansargues", "Saint-Just", "Mudaison", "Candillargues", "Saturargues",
            "Vauvert", "Aimargues", "Gallargues-le-Montueux", "Vergeze",
            "Bernis", "Milhaud", "Bouillargues", "Redessan", "Marguerittes",
            "Caveirac", "Saint-Gervasy", "Gajan", "Clarensac", "Langlade",
            "Calvisson", "Junas", "Villevieille", "Sommières", "Quissac",
            "Sauve", "Ganges", "Saint-Guilhem-le-Desert", "Gignac", "Aniane",
            "Clermont-l'Herault", "Montarnaud", "Saint-Andre-de-Sangonis",
            "Gigean", "Fabrègues", "Mireval", "Vic-la-Gardiole", "Frontignan",
            "Balaruc-les-Bains", "Balaruc-le-Vieux", "Bouzigues", "Loupian",
            "Meze", "Florensac", "Pinet", "Pomerols", "Marseillan", "Vias",
            "Portiragnes", "Valras-Plage", "Vendres", "Lespignan", "Serignan",
            "Sauvian", "Villeneuve-les-Beziers", "Montady", "Lieuran-les-Beziers",
            "Servian", "Roujan", "Gabian", "Pouzolles", "Alignan-du-Vent",
            "Montblanc", "Pinet", "Béziers", "Murviel-les-Beziers",
            "Lignan-sur-Orb", "Cessenon-sur-Orb", "Roquebrun", "Olargues",
            "Saint-Pons-de-Thomieres", "Lamalou-les-Bains", "Bedarieux",
            "Lunas", "Joncels", "Avene", "Octon", "Salasc", "Moureze",
            "Ceyras", "Paulhan", "Canet", "Aspiran", "Nizas", "Lezignan-la-Cebe",
            "Pomerols", "Pinet", "Portiragnes", "Agde", "Cap-d-Agde",
            "Royan", "Palavas", "Carnon", "Ange", "Fabrègues"
        ]
    },
    "Nouvelle-Aquitaine": {
        "villes": [
            "Bordeaux", "Limoges", "Poitiers", "La Rochelle", "Bayonne", "Pau",
            "Merignac", "Pessac", "Talence", "Angouleme", "Niort", "Brive-la-Gaillarde",
            "Perigueux", "Agen", "Mont-de-Marsan", "Saintes", "Cognac", "Rochefort",
            "Arcachon", "Cestas", "Le Bouscat", "Bruges", "Eysines", "Blanquefort",
            "Parempuyre", "Saint-Medard-en-Jalles", "Le Haillan", "Martignas-sur-Jalle",
            "Saint-Jean-d'Illac", "Le Taillan-Medoc", "Ludon-Medoc", "Macau",
            "Cantenac", "Margaux", "Lamarque", "Cussac-Fort-Medoc", "Pauillac",
            "Saint-Estephe", "Lesparre-Medoc", "Soulac-sur-Mer",
            "Ambares-et-Lagrave", "Carbon-Blanc", "Cenon", "Floirac", "Lormont",
            "Bassens", "Saint-Vincent-de-Paul", "Montussan", "Tresses",
            "Salleboeuf", "Bonnetan", "Camblanes-et-Meynac", "Carignan-de-Bordeaux",
            "Quinsac", "Latresne", "Bouliac", "Artigues-pres-Bordeaux",
            "Beychac-et-Caillau", "Sainte-Eulalie", "Saint-Loubès",
            "Saint-Sulpice-et-Cameyrac", "Pompignac", "Canejan", "Gradignan",
            "Villenave-d'Ornon", "Leognan", "Martillac", "Portets", "Podensac",
            "Cerons", "Barsac", "Preignac", "Langon", "Saint-Macaire",
            "La Reole", "Monségur", "Duras", "Miramont-de-Guyenne",
            "Marmande", "Tonneins", "Aiguillon", "Villeneuve-sur-Lot",
            "Fumel", "Penne-d'Agenais", "Casseneuil", "Cancon",
            "Sauveterre-la-Lemance", "Montflanquin", "Beauville", "Puymirol",
            "Astaffort", "Layrac", "Bon-Encontre", "Boé", "Foulayronnes",
            "Pont-du-Casse", "Roquefort", "Villeneuve-de-Marsan",
            "Hagetmau", "Dax", "Soustons", "Capbreton", "Hossegor", "Soorts-Hossegor",
            "Ondres", "Tarnos", "Anglet", "Biarritz", "Bidart", "Guethary",
            "Saint-Jean-de-Luz", "Ciboure", "Hendaye", "Urrugne", "Biriatou",
            "Mouguerre", "Saint-Pierre-d'Irube", "Lahonce", "Urt", "Guiche",
            "Bidache", "Came", "Hastingues", "Peyrehorade", "Sorde-l-Abbaye",
            "Orthevielle", "Saint-Lon-les-Mines", "Pouillon", "Amou",
            "Mugron", "Montfort-en-Chalosse", "Aire-sur-l-Adour", "Grenade-sur-l-Adour",
            "Tartas", "Sabres", "Morcenx", "Mimizan", "Biscarrosse",
            "Parentis-en-Born", "Gastes", "Sanguinet", "La Teste-de-Buch",
            "Gujan-Mestras", "Le Teich", "Biganos", "Audenge", "Lanton",
            "Andernos-les-Bains", "Lege-Cap-Ferret", "Arès", "Claouey",
            "Lanton", "Marcheprime", "Mios", "Salles"
        ]
    },
    "Hauts-de-France": {
        "villes": [
            "Lille", "Amiens", "Roubaix", "Tourcoing", "Dunkerque", "Arras",
            "Valenciennes", "Calais", "Boulogne-sur-Mer", "Lens", "Douai",
            "Villeneuve-d'Ascq", "Beauvais", "Compiegne", "Creil", "Soissons",
            "Laon", "Saint-Quentin", "Maubeuge", "Cambrai", "Bethune",
            "Henin-Beaumont", "Liévin", "Bruay-la-Buissière", "Noeux-les-Mines",
            "Carvin", "Libercourt", "Leforest", "Montigny-en-Ostrevent",
            "Pecquencourt", "Aniche", "Cambrai", "Denain", "Condé-sur-l-Escaut",
            "Valenciennes", "Anzin", "Beuvrages", "Bruay-sur-l-Escaut",
            "Fresnes-sur-Escaut", "Vieux-Condé", "Saint-Saulve",
            "Marly", "Aulnoy-lez-Valenciennes", "Sebourg", "Quarouble",
            "Jenlain", "Villers-Pol", "Roisin", "Bavay", "Bellignies",
            "Bermeries", "Mecquignies", "Gommegnies", "Poix-du-Nord",
            "Le Quesnoy", "Solesmes", "Avesnes-sur-Helpe", "Fourmies",
            "Wignehies", "Hirson", "Guise", "La Capelle", "Vervins",
            "Laon", "Chauny", "Tergnier", "La Fere", "Barisis-aux-Bois",
            "Amifontaine", "Berry-au-Bac", "Pontavert", "Condé-sur-Aisne",
            "Vailly-sur-Aisne", "Vic-sur-Aisne", "Soissons",
            "Compiègne", "Senlis", "Chantilly", "Creil", "Nogent-sur-Oise",
            "Montataire", "Liancourt", "Clermont", "Beauvais",
            "Breteuil", "Grandvilliers", "Formerie", "Crèvecoeur-le-Grand",
            "Marseille-en-Beauvaisis", "Saint-Just-en-Chaussée",
            "Clermont", "Estrées-Saint-Denis", "Verberie",
            "Pont-Sainte-Maxence", "Gouvieux", "Lamorlaye",
            "Chantilly", "Coye-la-Foret", "Luzarches", "Louvres",
            "Gonesse", "Goussainville", "Villeron", "Fosses",
            "Saint-Brice-sous-Foret", "Montmorency", "Taverny",
            "Saint-Leu-la-Foret", "Bessancourt", "Mery-sur-Oise"
        ]
    },
    "Grand Est": {
        "villes": [
            "Strasbourg", "Reims", "Metz", "Nancy", "Mulhouse", "Troyes",
            "Colmar", "Charleville-Mezieres", "Chaumont", "Bar-le-Duc",
            "Chalons-en-Champagne", "Epinal", "Saint-Dié-des-Vosges",
            "Thionville", "Forbach", "Sarreguemines", "Saint-Avold",
            "Haguenau", "Schiltigheim", "Illkirch-Graffenstaden",
            "Lingolsheim", "Oberhausbergen", "Ostwald", "Geispolsheim",
            "Eschau", "Plobsheim", "Innenheim", "Obernai", "Molsheim",
            "Mutzig", "Rosheim", "Barr", "Dambach-la-Ville", "Kintzheim",
            "Selestat", "Ribeauville", "Riquewihr", "Kaysersberg",
            "Ammerschwihr", "Katzenthal", "Ingersheim", "Wintzenheim",
            "Turckheim", "Zimmerbach", "Walbach", "Gunsbach",
            "Munster", "Soultzeren", "Sondernach", "Metzeral",
            "Stosswihr", "Mittlach", "Wildenstein", "Fellering",
            "Oderen", "Kruth", "Wesserling", "Saint-Amarin",
            "Masevaux", "Sewen", "Niederbruck", "Wegscheid",
            "Thann", "Wittelsheim", "Wittenheim", "Pfastatt",
            "Kingersheim", "Illzach", "Riedisheim", "Rixheim",
            "Habsheim", "Sierentz", "Bartenheim", "Saint-Louis",
            "Huningue", "Village-Neuf", "Chalampé", "Ottmarsheim",
            "Ensisheim", "Reguisheim", "Rumersheim-le-Haut",
            "Bollwiller", "Pulversheim", "Feldkirch", "Reiningue",
            "Lutterbach", "Morschwiller-le-Bas", "Zimmersheim",
            "Landser", "Hirsingue", "Altkirch", "Ferrette", "Delle",
            "Belfort", "Danjoutin", "Bavilliers", "Offemont",
            "Sevenans", "Trévenans", "Bourogne", "Valdoie",
            "Beaucourt", "Delle", "Grandvillars", "Morvillars",
            "Saint-Germain-le-Châtelet", "Rougemont-le-Château",
            "Giromagny", "Lepuix", "Auxelles-Bas", "Auxelles-Haut",
            "Chaux", "Etueffont", "Ronchamp", "Champagney",
            "Hericourt", "Montbeliard", "Audincourt", "Valentigney",
            "Pont-de-Roide", "Baume-les-Dames", "Belfontaine",
            "Besancon", "Pontarlier", "Mouthe", "Levier",
            "Salins-les-Bains", "Arbois", "Poligny", "Lons-le-Saunier",
            "Champagnole", "Morez", "Saint-Claude", "Oyonnax",
            "Nantua", "Bellegarde-sur-Valserine", "Culoz",
            "Seyssel", "Frangy", "Rumilly", "Alby-sur-Cheran"
        ]
    },
    "Pays de la Loire": {
        "villes": [
            "Nantes", "Angers", "Le Mans", "Saint-Nazaire", "Cholet", "La Roche-sur-Yon",
            "Saint-Herblain", "Laval", "Fontenay-le-Comte", "Les Sables-d'Olonne",
            "Rezé", "Coueron", "Saint-Sébastien-sur-Loire", "Vertou", "Bouguenais",
            "Orvault", "Sainte-Luce-sur-Loire", "Carquefou", "Treillieres",
            "Saint-Aignan-Grandlieu", "La Chevroliere", "Saint-Philbert-de-Grand-Lieu",
            "Machecoul", "Pornic", "Paimboeuf", "Saint-Brevin-les-Pins",
            "Corsept", "Frossay", "Saint-Viaud", "Vue", "Bouaye",
            "Pont-Saint-Martin", "Montbert", "Vieillevigne", "La Legere",
            "Corcoue-sur-Logne", "Saint-Lumine-de-Coutais", "Paulx",
            "Saint-Mars-de-Coutais", "Rouans", "La Montagne", "Indre",
            "Coueron", "Le Temple-de-Bretagne", "Grandchamp-des-Fontaines",
            "Treillières", "Casson", "Petit-Mars", "Carquefou",
            "Mauves-sur-Loire", "Thouaré-sur-Loire", "Sainte-Luce-sur-Loire",
            "La Chapelle-sur-Erdre", "Nort-sur-Erdre", "Héric",
            "Sucé-sur-Erdre", "Les Touches", "Nozay", "Blain", "Derval",
            "Châteaubriant", "Nozay", "Guemene-Penfao", "Redon",
            "Pontchâteau", "Saint-Gildas-des-Bois", "Herbignac",
            "La Baule-Escoublac", "Guerande", "Le Pouliguen",
            "Batz-sur-Mer", "Le Croisic", "Saint-Molf",
            "Asserac", "Mesquer", "Piriac-sur-Mer", "La Turballe",
            "Savenay", "Malville", "Cordemais", "Donges",
            "Montoir-de-Bretagne", "Saint-Nazaire", "Trignac",
            "La Chapelle-des-Marais", "Saint-Joachim", "Crossac"
        ]
    },
    "Bretagne": {
        "villes": [
            "Rennes", "Brest", "Quimper", "Vannes", "Saint-Malo", "Lorient",
            "Lannion", "Lanester", "Quimperle", "Concarneau", "Douarnenez",
            "Morlaix", "Landerneau", "Saint-Brieuc", "Dinan", "Fougeres",
            "Vitré", "Pontivy", "Auray", "Ploermel", "Guingamp",
            "Cesson-Sevigne", "Betton", "Saint-Gregoire", "Pacé", "Chavagne",
            "Bruz", "Chantepie", "Vezin-le-Coquet", "Le Rheu",
            "Thorigne-Fouillard", "Acigne", "Noyal-Chatillon-sur-Seiche",
            "Orgeres", "Pont-Pean", "Guichen", "Bain-de-Bretagne",
            "Crevin", "Laille", "Bourgbarré", "Janze", "Piré-sur-Seiche",
            "Acigne", "Chateaugiron", "Domloup", "Corps-Nuds",
            "Noyal-sur-Vilaine", "Vern-sur-Seiche", "La Mézière",
            "Melesse", "Parthenay-de-Bretagne", "Geveze",
            "Saint-Gilles", "Montgermont", "Monterfil",
            "Mordelles", "Goven", "Pleumeleuc", "Montfort-sur-Meu",
            "Bedee", "Saint-Meen-le-Grand", "Merdrignac",
            "Loudéac", "La Motte", "Saint-Caradec", "Plemet",
            "Mur-de-Bretagne", "Caurel", "Gouarec", "Rostrenen",
            "Carhaix-Plouguer", "Pleyben", "Chateaulin", "Pont-de-Buis-les-Quimerch",
            "Crozon", "Camaret-sur-Mer", "Landévennec", "Argol",
            "Telgruc-sur-Mer", "Ploeven", "Locronan", "Plogonnec",
            "Guengat", "Ergue-Gaberic", "Rosporden", "Scaer",
            "Bannalec", "Arzano", "Moelan-sur-Mer", "Clohars-Carnoet"
        ]
    },
    "Normandie": {
        "villes": [
            "Rouen", "Caen", "Le Havre", "Cherbourg", "Evreux", "Alencon",
            "Dieppe", "Saint-Lo", "Granville", "Argentan", "Vire",
            "Flers", "Lisieux", "Bayeux", "Coutances", "Avranches",
            "Mont-Saint-Michel", "Honfleur", "Deauville", "Trouville-sur-Mer",
            "Cabourg", "Dives-sur-Mer", "Houlgate", "Villers-sur-Mer",
            "Tourgeville", "Saint-Arnoult", "Benerville-sur-Mer",
            "Blonville-sur-Mer", "Auberville", "Gonneville-sur-Mer",
            "Dozule", "Dozulé", "Hotot-en-Auge", "Mezidon-Vallée-d'Auge",
            "Saint-Pierre-sur-Dives", "Livarot", "Orbec",
            "Bernay", "Pont-Audemer", "Pont-l'Evêque", "Beuzeville",
            "Quillebeuf-sur-Seine", "Tancarville", "Saint-Romain-de-Colbosc",
            "Gonfreville-l'Orcher", "Montivilliers", "Harfleur",
            "Sainte-Adresse", "Octeville-sur-Mer", "Epouville",
            "Gainneville", "Rogerville", "Gonfreville-l'Orcher",
            "Notre-Dame-de-Gravenchon", "Port-Jerome-sur-Seine",
            "Lillebonne", "Bolbec", "Fauville-en-Caux",
            "Goderville", "Criquetot-l'Esneval", "Etretat",
            "Fecamp", "Valmont", "Veulettes-sur-Mer",
            "Saint-Valery-en-Caux", "Cany-Barville",
            "Doudeville", "Yvetot", "Maromme", "Barentin",
            "Pavilly", "Deville-les-Rouen", "Bois-Guillaume",
            "Bihorel", "Montville", "Cleres", "Buchy",
            "Forges-les-Eaux", "Gournay-en-Bray", "Argueil"
        ]
    },
    "Bourgogne-Franche-Comte": {
        "villes": [
            "Dijon", "Besancon", "Chalon-sur-Saone", "Nevers", "Montbeliard", "Macon",
            "Auxerre", "Sens", "Belfort", "Vesoul", "Lons-le-Saunier",
            "Pontarlier", "Beaune", "Autun", "Le Creusot", "Montceau-les-Mines",
            "Gueugnon", "Paray-le-Monial", "Charolles", "Bourbon-Lancy",
            "Digoin", "Clamecy", "Decize", "Cosne-Cours-sur-Loire",
            "Varzy", "Pougues-les-Eaux", "Saint-Benin-d'Azy",
            "Imphy", "Fourchambault", "Guerigny",
            "Cercy-la-Tour", "Luzy", "Fours",
            "Toulon-sur-Arroux", "Issy-l'Eveque", "Chateau-Chinon",
            "Lormes", "Corbigny", "Avallon", "Vezelay",
            "Noyers", "Chablis", "Tonnerre", "Migennes",
            "Joigny", "Villeneuve-sur-Yonne", "Sens",
            "Nemours", "Montereau-Fault-Yonne",
            "Melun", "Fontainebleau", "Moret-Loing-et-Orvanne",
            "Champagne-sur-Seine", "Bray-sur-Seine", "Provins"
        ]
    },
    "Centre-Val de Loire": {
        "villes": [
            "Tours", "Orleans", "Chartres", "Blois", "Chateauroux", "Bourges",
            "Vierzon", "Romorantin-Lanthenay", "Dreux", "Vendome",
            "Amboise", "Chinon", "Loches", "Joué-lès-Tours", "Saint-Cyr-sur-Loire",
            "La Riche", "Saint-Pierre-des-Corps", "Saint-Average",
            "Chambray-les-Tours", "Ballan-Mire", "Savigne-sur-Lathan",
            "Langeais", "Cinq-Mars-la-Pile", "Luynes", "Fondettes",
            "Neuille-Pont-Pierre", "Chateau-la-Valliere", "Neuvy-le-Roi",
            "Chateau-Renault", "Amboise", "Nazelles-Negron",
            "Pocé-sur-Cisse", "Noizay", "Vernou-sur-Brenne",
            "Vouvray", "Sainte-Radegonde", "Rochecorbon",
            "Parcay-Meslay", "Chanceaux-sur-Choisille",
            "Avrille-les-Ponceaux", "Continvoir", "Channay-sur-Lathan",
            "Savigny-en-Veron", "Avoine", "Beaumont-en-Veron",
            "Candes-Saint-Martin", "Montsoreau",
            "Saumur", "Doue-en-Anjou", "Montreuil-Bellay",
            "Thouars", "Bressuire",
            "Gien", "Montargis", "Pithiviers",
            "Etampes", "Angerville", "Milly-la-Foret",
            "Nemours", "Fontainebleau"
        ]
    },
    "Corse": {
        "villes": [
            "Ajaccio", "Bastia", "Corte", "Calvi", "Porto-Vecchio",
            "Bonifacio", "Sartene", "Propriano", "Ile-Rousse",
            "Calenzana", "Balagne", "Belgodere", "Algajola",
            "Lumio", "Avapessa", "Montegrosso", "Aregno",
            "Lavatoggio", "Cateri", "Sant'Antonino", "Feliceto",
            "Muro", "Speloncato", "Occhiatana", "Novella",
            "Pietralba", "Palasca", "Costa", "Lama", "Urtaca",
            "Poggio-Mezzana", "Moïta", "Quercitellu", "Cervione",
            "Prunete", "Moriani-Plage", "San-Nicolao",
            "Antisanti", "Linguizzetta", "Tallone",
            "Ghisonaccia", "Aghione", "Prunelli-di-Fiumorbu",
            "Serra-di-Fiumorbu", "Solenzara", "Zonza",
            "Levie", "Santa-Lucia-di-Tallano",
            "Figari", "Pianottoli-Caldarello",
            "Monacia-d'Aullene", "Sotta",
            "Santa-Lucia-di-Porto-Vecchio"
        ]
    },
    "Dom-Tom": {
        "villes": [
            "Saint-Denis", "Saint-Paul", "Le Tampon", "Saint-Pierre", "Saint-Andre",
            "Fort-de-France", "Le Lamentin", "Le Robert", "Sainte-Marie", "Schoelcher",
            "Pointe-a-Pitre", "Baie-Mahault", "Le Gosier", "Abymes", "Sainte-Anne",
            "Cayenne", "Saint-Laurent-du-Maroni", "Kourou", "Matoury", "Remire-Montjoly",
            "Marigot", "Gustavia", "Dzaoudzi", "Mamoudzou", "Koungou",
            "Bandraboua", "Boueni", "Kani-Keli"
        ]
    }
}
def _find_region_ville(location: str) -> tuple[str, str]:
    """
    Cherche la région et la ville à partir du champ location.
    Retourne (region, ville) ou ("", "") si non trouvé.
    """
    if not location or location.lower() in ("", "nan", "none"):
        return "", ""
    
    location_lower = location.lower()

    # Nettoyer les mots inutiles
    for word in ["greater", "area", "metropolitan", "agglomération", "région de"]:
        location_lower = location_lower.replace(word.lower(), "")

    def normalize(s: str) -> str:
        return s.lower().replace("-", "").replace(" ", "").replace("'", "").replace("'", "")

    location_norm = normalize(location_lower)

    # Chercher dans les villes et régions
    for region, data in regions_villes.items():
        for ville in data["villes"]:
            if normalize(ville) in location_norm:
                return region, ville

    return "", ""


def Rephrase(db: Session, base: str = "leads"):
    """
    Reformule le champ location avec le format: City, Region, Country
    Exemple: "Paris, Ile-de-France, France"
    """
    try:
        if "silver" in base.lower():
            leads = db.query(Leads).all()
        else:
            leads = db.query(Leads).all()
        
        updated_count = 0
        
        for lead in leads:
            if not lead.location or lead.location.lower() in ("", "nan", "none"):
                continue

            region, ville = _find_region_ville(lead.location)

            if region and ville:
                new_location = f"{ville}, {region}, France"
            else:
                parts = lead.location.strip().split()
                if len(parts) == 3:
                    new_location = f"{parts[0]}, {parts[1]}, {parts[2]}"
                else:
                    continue

            if lead.location != new_location:
                lead.location = new_location
                updated_count += 1
        
        db.commit()
        
        print(f"✅ {updated_count} locations reformulées")
        return {"reformulated": updated_count}
    
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur base de données : {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inattendue : {str(e)}")




# ---------------------------------------------------------------------------
# Vérification d'email par sonde SMTP RCPT (sans envoyer d'email)
# ---------------------------------------------------------------------------
_MX_CACHE: dict = {}          # domaine -> [hosts MX] (ou None si aucun)
_CATCHALL_CACHE: dict = {}    # domaine -> bool (accept-all)
_MX_LOCK = threading.Lock()


def _sender_domain() -> str:
    u = os.getenv("SMTP_HELO_DOMAIN") or os.getenv("SMTP_USER") or "example.com"
    return u.split("@")[-1].strip() or "example.com"


def _resolve_mx(domain: str):
    """Retourne la liste des hosts MX (par priorité), fallback A record. Caché."""
    with _MX_LOCK:
        if domain in _MX_CACHE:
            return _MX_CACHE[domain]
    hosts = None
    try:
        recs = dns.resolver.resolve(domain, "MX")
        hosts = [str(r.exchange).rstrip(".") for r in sorted(recs, key=lambda r: r.preference)]
    except Exception:
        try:
            dns.resolver.resolve(domain, "A")
            hosts = [domain]   # pas de MX -> le domaine lui-même
        except Exception:
            hosts = None
    with _MX_LOCK:
        _MX_CACHE[domain] = hosts
    return hosts


def _rcpt_code(mx_host: str, sender_domain: str, target: str, timeout: int) -> int:
    """Ouvre une session SMTP vers le MX et renvoie le code RCPT TO (ou 0 si injoignable)."""
    try:
        with smtplib.SMTP(mx_host, 25, timeout=timeout) as smtp:
            smtp.ehlo(sender_domain)
            smtp.mail(f"probe@{sender_domain}")
            code, _ = smtp.rcpt(target)
            smtp.quit()
            return int(code)
    except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError, socket.error, OSError):
        return 0
    except smtplib.SMTPException:
        return 0


def _is_catch_all(mx_host: str, sender_domain: str, domain: str, timeout: int) -> bool:
    """Le domaine accepte-t-il n'importe quelle adresse ? (caché par domaine)"""
    with _MX_LOCK:
        if domain in _CATCHALL_CACHE:
            return _CATCHALL_CACHE[domain]
    probe = f"nonexistent-{uuid.uuid4().hex[:12]}@{domain}"
    code = _rcpt_code(mx_host, sender_domain, probe, timeout)
    catchall = code in (250, 251)
    with _MX_LOCK:
        _CATCHALL_CACHE[domain] = catchall
    return catchall


def smtp_probe(email_addr: str, timeout: int = 8) -> dict:
    """
    Vérifie une adresse via RCPT TO (aucun email envoyé).
    code: 250 = valide/livrable, 550 = invalide, 450 = inconnu (greylist/catch-all/injoignable).
    """
    addr = str(email_addr or "").strip().lower()
    if "@" not in addr:
        return {"email": addr, "code": 550, "status": "❌ format invalide"}
    domain = addr.split("@", 1)[1]
    hosts = _resolve_mx(domain)
    if not hosts:
        return {"email": addr, "code": 450, "status": "⚠️ domaine sans MX"}

    sender = _sender_domain()
    for mx in hosts[:2]:   # essaie les 2 premiers MX
        code = _rcpt_code(mx, sender, addr, timeout)
        if code == 0:
            continue       # MX injoignable -> essayer le suivant
        if code in (250, 251):
            # Domaine accept-all ? -> on ne peut pas conclure "valide"
            if _is_catch_all(mx, sender, domain, timeout):
                return {"email": addr, "code": 450, "status": "⚠️ domaine catch-all"}
            return {"email": addr, "code": 250, "status": "✅ adresse acceptée"}
        if code in (550, 551, 553, 554):
            return {"email": addr, "code": 550, "status": "❌ adresse rejetée"}
        if 400 <= code < 500:
            return {"email": addr, "code": 450, "status": "⚠️ temporaire (greylist)"}
    return {"email": addr, "code": 450, "status": "⚠️ MX injoignable"}


def _apply_statu(db: Session, to_email: str, statu: str):
    """Met à jour le statut sur silver/gold (par email) + tous les leads applique correspondants."""
    lead = db.query(Leads).filter(Leads.email == to_email).first()
    if not lead:
        lead = db.query(Leads).filter(Leads.email == to_email).first()
    if lead:
        lead.statu = statu
    db.query(SteagingApplique).filter(SteagingApplique.email == to_email).update(
        {SteagingApplique.statu: statu}, synchronize_session=False
    )
    db.commit()


EMAIL_OBJET = "Proposition Bureau #05202404-10432345"

EMAIL_CORPS = """Bonjour,

Représentant de LYNK FREETEK, nous disposons d'une équipe compétente pour vous accompagner sur les recherches bureaux commerciales sur toute la France

J'aimerais discuter des possibilités de partenariats lors d'une visioconférence ou d'un appel.

Merci de me faire part de vos disponibilités.


Restant à votre disposition pour plus d'informations.


Service commercial
"""


def send_email(to_email: str, message_id: str) -> bool:
    """
    Envoie l'email à l'adresse donnée.
    Lève une exception smtplib si le relais refuse le destinataire.
    """

    SMTP_USER= os.getenv("SMTP_USER")
    SMTP_HOST=os.getenv("SMTP_HOST")
    SMTP_PORT=int(os.getenv("SMTP_PORT") or 587)
    SMTP_PASSWORD=os.getenv("SMTP_PASSWORD")

    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = to_email
    msg["Subject"] = EMAIL_OBJET
    # Indispensable : check_bounce recherche ce Message-ID dans les retours
    # mailer-daemon. Sans lui, on ne peut pas relier un bounce à son envoi.
    msg["Message-ID"] = message_id
    # charset utf-8 obligatoire : le corps contient des accents (« Représentant »,
    # « disponibilités »). Sans ça, MIMEText part en us-ascii et les casse.
    msg.attach(MIMEText(EMAIL_CORPS, "plain", "utf-8"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.sendmail(SMTP_USER, to_email, msg.as_string())
    return True


def check_bounce(message_id: str, to_email: str, limite: int = 10) -> dict | None:
    """Cherche un retour mailer-daemon correspondant à cet envoi.

    `limite` = nombre de derniers bounces inspectés. Pour un lot d'envois, il
    faut l'élargir : sinon les bounces au-delà des N derniers sont manqués et
    l'adresse serait déclarée valide à tort.
    """
    IMAP_HOST     = os.getenv("IMAP_HOST")
    SMTP_USER     = os.getenv("SMTP_USER")
    SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

    with imaplib.IMAP4_SSL(IMAP_HOST) as imap:
        imap.login(SMTP_USER, SMTP_PASSWORD)
        imap.select("INBOX")

        _, data = imap.search(None, 'OR FROM "mailer-daemon" FROM "postmaster"')
        ids = data[0].split()

        for mail_id in reversed(ids[-limite:]):
            _, msg_data = imap.fetch(mail_id, "(RFC822)")
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            # Récupère tout le contenu du mail (headers + body)
            full_content = raw.decode(errors="ignore").lower()

            # Cherche soit le message_id soit l'adresse email directement
            if message_id.lower() in full_content or to_email.lower() in full_content:
                # Extrait le body
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            body += part.get_payload(decode=True).decode(errors="ignore")
                else:
                    body = msg.get_payload(decode=True).decode(errors="ignore")

                return {"subject": msg["subject"], "body": body[:300]}

    return None
 
 
def send_and_check(to_email: str, db: Session = None, attente: int = 20) -> dict:
    """
    Envoie un vrai email de test à l'adresse, puis attend un éventuel retour
    mailer-daemon (bounce) pendant `attente` secondes pour conclure.

    250 -> aucun bounce reçu (adresse considérée valide)
    550 -> destinataire refusé à l'envoi, ou bounce reçu
    450 -> indéterminé (erreur SMTP/IMAP)

    ⚠️ L'absence de bounce dans la fenêtre d'attente NE PROUVE PAS que la boîte
    existe : un rejet peut arriver plusieurs heures après. Le verdict 250 est
    donc optimiste. La sonde smtp_probe() reste disponible pour un test
    immédiat et sans envoi.
    """
    addr = str(to_email or "").strip().lower()
    if "@" not in addr:
        return {"email": addr, "status": "❌ format invalide", "code": 550, "raison": "❌ format invalide"}

    message_id = make_msgid()

    # 1) Envoi. Un refus immédiat du destinataire tranche déjà la question.
    try:
        send_email(addr, message_id)
    except smtplib.SMTPRecipientsRefused:
        res = {"email": addr, "code": 550, "status": "❌ destinataire refusé à l'envoi"}
    except Exception as e:
        res = {"email": addr, "code": 450, "status": f"⚠️ envoi impossible ({e})"}
    else:
        # 2) Attente d'un bounce : on interroge la boîte toutes les 5 s.
        bounce = None
        fin = time.time() + max(0, attente)
        while time.time() < fin:
            time.sleep(5)
            try:
                bounce = check_bounce(message_id, addr)
            except Exception as e:
                print(f"⚠️ Lecture IMAP impossible pour {addr} : {e}")
                break
            if bounce:
                break
        if bounce:
            res = {"email": addr, "code": 550, "status": f"❌ rejeté : {bounce.get('subject') or 'bounce reçu'}"}
        else:
            res = {"email": addr, "code": 250, "status": f"✅ aucun rejet en {attente}s"}

    is_valid = int(res.get("code", 0) or 0) == 250
    statu = "disponible" if is_valid else "non disponible"
    if db:
        _apply_statu(db, addr, statu)
    return {
        "email":  addr,
        "status": res.get("status", ""),
        "code":   res.get("code", 0),
        "raison": res.get("status", ""),
    }


def send_and_check_bulk(emails: list, db: Session = None, attente: int = 20) -> list:
    """
    Envoie un email de test à chaque adresse, puis attend UNE SEULE fois avant
    de relever les bounces.

    Appeler send_and_check() en boucle multiplierait l'attente par le nombre
    d'adresses (20s x N) et ferait expirer la requête.
    """
    envois = []
    for brut in emails or []:
        addr = str(brut or "").strip().lower()
        if "@" not in addr:
            envois.append((addr, None, {"code": 550, "status": "❌ format invalide"}))
            continue
        message_id = make_msgid()
        try:
            send_email(addr, message_id)
            envois.append((addr, message_id, None))
        except smtplib.SMTPRecipientsRefused:
            envois.append((addr, None, {"code": 550, "status": "❌ destinataire refusé à l'envoi"}))
        except Exception as e:
            envois.append((addr, None, {"code": 450, "status": f"⚠️ envoi impossible ({e})"}))

    # Une seule fenêtre d'attente, commune à tout le lot
    if any(mid for _, mid, _ in envois):
        time.sleep(max(0, attente))

    resultats = []
    for addr, message_id, verdict in envois:
        if verdict is None:
            try:
                # Fenêtre de lecture élargie : un lot génère plusieurs bounces
                bounce = check_bounce(message_id, addr, limite=max(10, 4 * len(envois)))
            except Exception as e:
                print(f"⚠️ Lecture IMAP impossible pour {addr} : {e}")
                bounce = None
            verdict = (
                {"code": 550, "status": f"❌ rejeté : {bounce.get('subject') or 'bounce reçu'}"}
                if bounce
                else {"code": 250, "status": f"✅ aucun rejet en {attente}s"}
            )
        statu = "disponible" if int(verdict.get("code", 0) or 0) == 250 else "non disponible"
        if db:
            _apply_statu(db, addr, statu)
        resultats.append({
            "email":  addr,
            "status": verdict.get("status", ""),
            "code":   verdict.get("code", 0),
            "raison": verdict.get("status", ""),
        })
    return resultats


# ---------------------------------------------------------------------------
# Vérification en masse — tâche de fond concurrente avec suivi de progression
# ---------------------------------------------------------------------------
VERIFY_JOBS: dict = {}
_JOBS_LOCK = threading.Lock()
_VERIFY_WORKERS = 100


def _verify_lead_task(lead_id: int, company_map: dict) -> str:
    """Vérifie un lead applique dans sa propre session ; renvoie le statut final."""
    db = SessionLocal()
    try:
        lead = db.query(SteagingApplique).filter(SteagingApplique.id == lead_id).first()
        if not lead:
            return "introuvable"
        res = _verify_one_applique(db, lead, dict(company_map))
        return res.get("statu", "non disponible")
    except Exception:
        db.rollback()
        return "erreur"
    finally:
        db.close()


def _run_verify_job(job_id: str, ids: list):
    # snapshot du map société (lecture unique)
    db0 = SessionLocal()
    try:
        company_map = _company_map_regex_patterne(db0)
    finally:
        db0.close()

    try:
        with ThreadPoolExecutor(max_workers=_VERIFY_WORKERS) as pool:
            futures = {pool.submit(_verify_lead_task, i, company_map): i for i in ids}
            for fut in as_completed(futures):
                statu = fut.result()
                with _JOBS_LOCK:
                    job = VERIFY_JOBS.get(job_id)
                    if not job:
                        continue
                    job["done"] += 1
                    if statu == "disponible":
                        job["disponible"] += 1
                    elif statu == "non disponible":
                        job["non_disponible"] += 1
                    else:
                        job["erreurs"] += 1
    finally:
        with _JOBS_LOCK:
            if job_id in VERIFY_JOBS:
                VERIFY_JOBS[job_id]["status"] = "done"


def start_verify_job(ids: list) -> dict:
    ids = [int(i) for i in (ids or []) if str(i).strip() != ""]
    job_id = uuid.uuid4().hex
    with _JOBS_LOCK:
        VERIFY_JOBS[job_id] = {
            "status": "running", "total": len(ids), "done": 0,
            "disponible": 0, "non_disponible": 0, "erreurs": 0,
        }
    if ids:
        threading.Thread(target=_run_verify_job, args=(job_id, ids), daemon=True).start()
    else:
        with _JOBS_LOCK:
            VERIFY_JOBS[job_id]["status"] = "done"
    return {"job_id": job_id, "total": len(ids)}


def get_verify_job(job_id: str) -> dict:
    with _JOBS_LOCK:
        job = VERIFY_JOBS.get(job_id)
        return dict(job) if job else {"status": "unknown"}


def _company_map_regex_patterne(db: Session):
    """map: nom société normalisé -> (regex, patterne)"""
    rows = db.query(societeleads.nom, societeleads.regex, societeleads.patterne).all()
    m = {}
    for nom, rgx, patt in rows:
        k = _norm_company_key(nom)
        if k:
            m[k] = (str(rgx or "").strip(), str(patt or "").strip())
    return m


def _patterne_to_regex(patterne: str) -> str:
    """
    Construit un regex générique à partir du patterne : on ne verrouille que le DOMAINE
    (partie après @) ; le local-part accepte tous les formats de nom possibles
    ({prenom}.{nom}, {p}.{nom}, {nom}.{prenom}, {p}{nom}, {n}{prenom}, jd, ...).
    Ex: "{prenom}.{nom}@soprat.fr" -> "^[a-z]+([._-][a-z]+)*@soprat\\.fr$"
    """
    pat = (patterne or "").strip()
    if "@" not in pat:
        return ""
    domain = pat.split("@", 1)[1].strip()
    if not domain:
        return ""
    return "^[a-z]+([._-][a-z]+)*@" + re.escape(domain) + "$"


def _split_lines(s) -> list:
    """Découpe une valeur multi-lignes en liste (une entrée par ligne non vide)."""
    return [x.strip() for x in str(s or "").replace("\r", "").split("\n") if x.strip()]


def _autoadd_societe_from_email(db: Session, nom_soc: str, email: str, prenom, nom):
    """
    À partir d'un email livré : dérive patterne + regex.
    - Société inexistante -> on la crée.
    - Société existante -> on AJOUTE ce format s'il est nouveau (multi-patternes).
    Retourne (regex, patterne) [valeurs stockées, multi-lignes] si créé/modifié, sinon None.
    """
    nom_soc = (nom_soc or "").strip()
    if not nom_soc:
        return None
    import service.serviceSociete as sso
    patterne = sso.derive_patterne(email, prenom, nom)
    if not patterne:
        return None
    regex = _patterne_to_regex(patterne)
    try:
        exists = db.query(societeleads).filter(societeleads.nom.ilike(nom_soc)).first()
        if not exists:
            db.add(societeleads(nom=nom_soc, patterne=patterne, regex=regex))
            db.commit()
            return (regex, patterne)
        # Société existante : ajouter le format s'il n'y est pas déjà
        patts = _split_lines(exists.patterne)
        if patterne in patts:
            return None
        patts.append(patterne)
        regs = _split_lines(exists.regex)
        if regex and regex not in regs:
            regs.append(regex)
        exists.patterne = "\n".join(patts)
        exists.regex = "\n".join(regs)
        db.commit()
        return (exists.regex, exists.patterne)
    except Exception:
        db.rollback()
        return None


def _verify_one_applique(db: Session, lead, company_map: dict, envoyer_test: bool = False) -> dict:
    """
    Flux de vérification :
    1) On confronte l'email du lead au patterne de sa société (via la regex stockée) :
       - conforme     -> on garde l'email du lead tel quel
       - non conforme -> on le régénère depuis le patterne
    2) On teste l'adresse retenue :
       - envoyer_test=True  -> envoi d'un VRAI email de test + attente d'un bounce
                               (250 = aucun rejet -> disponible)
       - envoyer_test=False -> sonde SMTP RCPT, aucun envoi (utilisé par la
                               vérification de masse : 100 envois simultanés
                               feraient suspendre le compte SMTP).
    """
    email = (lead.email or "").strip().lower()
    nom_soc = (lead.societe or "").strip()
    key = _norm_company_key(nom_soc)
    regex_raw, patterne_raw = company_map.get(key, ("", ""))
    patternes = _split_lines(patterne_raw)
    regexes = _split_lines(regex_raw)

    # 1) L'email du lead correspond-il au patterne de sa société ?
    conforme = False
    for rgx in regexes:
        try:
            if email and re.match(rgx, email, re.IGNORECASE):
                conforme = True
                break
        except re.error:
            continue

    cible = email
    regenere = False
    if not conforme and patternes:
        p = _norm_name_part(lead.prenom)
        n = _norm_name_part(lead.nom)
        for patt in patternes:
            gen = _build_email(patt, p, n)
            gen = (NettoyerUnEmail(gen) or gen or "").strip().lower()
            if gen and "@" in gen and "{" not in gen:
                cible = gen
                regenere = True
                break

    if not cible or "@" not in cible:
        lead.statu = "non disponible"
        db.commit()
        return {"id": lead.id, "email": cible, "statu": "non disponible",
                "regenerated": False, "conforme": conforme, "trusted": False}

    lead.email = cible
    db.commit()

    # 2) Test de l'adresse retenue
    if envoyer_test:
        res = send_and_check(cible, db)   # envoi réel + bounce ; met aussi statu à jour
        code = int(res.get("code", 0) or 0)
        statu = "disponible" if code == 250 else "non disponible"
    else:
        code = int(smtp_probe(cible).get("code", 0) or 0)
        # 450 = catch-all / greylist : non concluant, on ne condamne pas l'adresse
        statu = "disponible" if code in (250, 450) else "non disponible"
        lead.statu = statu
        db.commit()

    # L'email d'origine était conforme et accepté -> on apprend la société
    if statu == "disponible" and not regenere:
        added = _autoadd_societe_from_email(db, nom_soc, cible, lead.prenom, lead.nom)
        if added and key:
            company_map[key] = added

    return {"id": lead.id, "email": cible, "statu": statu,
            "regenerated": regenere, "conforme": conforme, "code": code,
            "trusted": False}


def VerifyAppliqueLead(db: Session, lead_id: int) -> dict:
    """Vérification à l'unité (bouton « Vérifier email ») : envoi d'un vrai
    email de test sur l'adresse retenue."""
    lead = db.query(SteagingApplique).filter(SteagingApplique.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead introuvable")
    cm = _company_map_regex_patterne(db)
    return _verify_one_applique(db, lead, cm, envoyer_test=True)


def VerifyAppliqueBulk(db: Session, ids: list) -> dict:
    if not ids:
        return {"results": [], "verified": 0}
    cm = _company_map_regex_patterne(db)
    leads = db.query(SteagingApplique).filter(SteagingApplique.id.in_(ids)).all()
    results = []
    for lead in leads:
        try:
            results.append(_verify_one_applique(db, lead, cm))
        except Exception as e:
            results.append({"id": lead.id, "statu": "erreur", "reason": str(e)})
    return {"results": results, "verified": len(results)}


def GenerateAppliqueEmail(db: Session, lead_id: int) -> dict:
    """
    Génère l'email d'un lead applique depuis le 1er patterne de sa société et le sauvegarde.
    PAS de vérification SMTP, PAS d'envoi vers Silver — simple remplissage.
    """
    lead = db.query(SteagingApplique).filter(SteagingApplique.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead introuvable")

    nom_soc = (lead.societe or "").strip()
    soc = db.query(societeleads).filter(societeleads.nom.ilike(nom_soc)).first() if nom_soc else None
    patternes = _split_lines(soc.patterne) if soc else []
    if not patternes:
        return {"id": lead.id, "email": lead.email, "error": "societe_inconnue"}

    gen = _build_email(patternes[0], _norm_name_part(lead.prenom), _norm_name_part(lead.nom))
    gen = (NettoyerUnEmail(gen) or gen or "").strip().lower()
    if not gen or "@" not in gen or "{" in gen:
        return {"id": lead.id, "email": lead.email, "error": "generation_impossible"}

    lead.email = gen
    db.commit()
    return {"id": lead.id, "email": gen}